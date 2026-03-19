import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── COLOR PALETTE ──
DARK_BG = "1B2A4A"
ACCENT_BG = "2E75B6"
LIGHT_BG = "D6E4F0"
WHITE = "FFFFFF"
HEADER_FONT = Font(name="Arial", bold=True, color=WHITE, size=11)
TITLE_FONT = Font(name="Arial", bold=True, color=WHITE, size=14)
LABEL_FONT = Font(name="Arial", size=10)
INPUT_FONT = Font(name="Arial", color="0000FF", size=10)  # Blue = hardcoded inputs
FORMULA_FONT = Font(name="Arial", color="000000", size=10)  # Black = formulas
PCT_FMT = '0.0%'
DOLLAR_FMT = '$#,##0'
DOLLAR_FMT_NEG = '$#,##0;($#,##0);"-"'
NUM_FMT = '#,##0'
thin_border = Border(
    left=Side(style='thin', color='B0B0B0'),
    right=Side(style='thin', color='B0B0B0'),
    top=Side(style='thin', color='B0B0B0'),
    bottom=Side(style='thin', color='B0B0B0')
)
input_fill = PatternFill('solid', fgColor='FFFFF0')
dark_fill = PatternFill('solid', fgColor=DARK_BG)
accent_fill = PatternFill('solid', fgColor=ACCENT_BG)
light_fill = PatternFill('solid', fgColor=LIGHT_BG)

def style_range(ws, row, col_start, col_end, font=None, fill=None, alignment=None, number_format=None, border=None):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        if font: cell.font = font
        if fill: cell.fill = fill
        if alignment: cell.alignment = alignment
        if number_format: cell.number_format = number_format
        if border: cell.border = border

# ════════════════════════════════════════
# SHEET 1: ASSUMPTIONS
# ════════════════════════════════════════
ws = wb.active
ws.title = "Assumptions"
ws.sheet_properties.tabColor = DARK_BG

for c in range(1, 10):
    ws.column_dimensions[get_column_letter(c)].width = [2, 35, 16, 16, 16, 16, 16, 16, 2][c-1]

# Title
ws.merge_cells('B1:G1')
ws['B1'] = "SURF BETTING SCENARIO MODEL"
ws['B1'].font = TITLE_FONT
ws['B1'].fill = dark_fill
style_range(ws, 1, 2, 7, fill=dark_fill)
ws.merge_cells('B2:G2')
ws['B2'] = "Back-of-Napkin Sizing | All figures illustrative"
ws['B2'].font = Font(name="Arial", italic=True, color="A0A0A0", size=9)
ws['B2'].fill = dark_fill
style_range(ws, 2, 2, 7, fill=dark_fill)

# ── PLATFORM ASSUMPTIONS ──
r = 4
ws.merge_cells(f'B{r}:G{r}')
ws[f'B{r}'] = "PLATFORM ASSUMPTIONS"
style_range(ws, r, 2, 7, font=HEADER_FONT, fill=accent_fill)

r = 5
headers = ["Platform", "Max Bet / Event", "Events / Year", "Surf Available?", "Notes"]
for i, h in enumerate(headers):
    ws.cell(row=r, column=2+i, value=h).font = Font(name="Arial", bold=True, size=9, color="333333")
    ws.cell(row=r, column=2+i).fill = light_fill
    ws.cell(row=r, column=2+i).border = thin_border
    ws.cell(row=r, column=2+i).alignment = Alignment(horizontal='center')

platforms = [
    ("DraftKings", 500, 11, "Yes", "US primary; state-dependent"),
    ("FanDuel", 500, 11, "Yes", "US primary; sharp-friendly"),
    ("BetMGM", 500, 11, "Yes", "Fast to limit winners"),
    ("Bet365", 500, 11, "Yes", "Gradual limiting"),
    ("Caesars", 500, 11, "Yes", "US; limited prop coverage"),
    ("BetUS", 500, 11, "Yes", "Offshore; higher limits possible"),
    ("TAB (Australia)", 500, 11, "Yes", "Deepest surf market"),
    ("Sportsbet (AU)", 500, 11, "Yes", "Australian market"),
]

for i, (name, bet, events, avail, notes) in enumerate(platforms):
    r = 6 + i
    ws.cell(row=r, column=2, value=name).font = LABEL_FONT
    ws.cell(row=r, column=3, value=bet).font = INPUT_FONT
    ws.cell(row=r, column=3).number_format = DOLLAR_FMT
    ws.cell(row=r, column=4, value=events).font = INPUT_FONT
    ws.cell(row=r, column=5, value=avail).font = LABEL_FONT
    ws.cell(row=r, column=5).alignment = Alignment(horizontal='center')
    ws.cell(row=r, column=6, value=notes).font = Font(name="Arial", size=9, italic=True, color="666666")
    for c in range(2, 7):
        ws.cell(row=r, column=c).border = thin_border
        ws.cell(row=r, column=c).fill = input_fill if c in [3, 4] else PatternFill()

PLAT_LAST = 13  # row 13 = last platform row (row 6+7)
r = PLAT_LAST + 1
ws.cell(row=r, column=2, value="Total Platforms").font = Font(name="Arial", bold=True, size=10)
ws.cell(row=r, column=3).font = FORMULA_FONT
ws[f'C{r}'] = f'=COUNTA(B6:B{PLAT_LAST})'
ws.cell(row=r, column=3).number_format = NUM_FMT
for c in range(2, 7):
    ws.cell(row=r, column=c).border = thin_border

# ── MODEL EDGE ASSUMPTIONS ──
r = PLAT_LAST + 3  # row 16
EDGE_START = r
ws.merge_cells(f'B{r}:G{r}')
ws[f'B{r}'] = "MODEL EDGE ASSUMPTIONS"
style_range(ws, r, 2, 7, font=HEADER_FONT, fill=accent_fill)

r += 1
edge_headers = ["Parameter", "Low (Conservative)", "Medium (Realistic)", "High (Optimistic)"]
for i, h in enumerate(edge_headers):
    ws.cell(row=r, column=2+i, value=h).font = Font(name="Arial", bold=True, size=9, color="333333")
    ws.cell(row=r, column=2+i).fill = light_fill
    ws.cell(row=r, column=2+i).border = thin_border

assumptions = [
    ("Model edge over market (% above break-even)", 0.02, 0.05, 0.10),
    ("Win rate on bets placed", 0.52, 0.55, 0.60),
    ("Avg odds on bets placed (decimal)", 2.0, 2.5, 3.0),
    ("Bets placed per event (across all platforms)", 3, 6, 10),
    ("Events bet on per year", 8, 10, 11),
    ("Average vig / juice", 0.0476, 0.0476, 0.0476),
    ("Platform utilization (% of platforms actually used)", 0.50, 0.75, 1.00),
]

for i, (label, low, mid, high) in enumerate(assumptions):
    r = EDGE_START + 2 + i
    ws.cell(row=r, column=2, value=label).font = LABEL_FONT
    for j, val in enumerate([low, mid, high]):
        cell = ws.cell(row=r, column=3+j, value=val)
        cell.font = INPUT_FONT
        cell.border = thin_border
        cell.fill = input_fill
        cell.alignment = Alignment(horizontal='center')
        if "%" in label or "vig" in label or "rate" in label or "utilization" in label:
            cell.number_format = PCT_FMT
        elif "odds" in label.lower():
            cell.number_format = '0.0'
        else:
            cell.number_format = NUM_FMT
    ws.cell(row=r, column=2).border = thin_border

# Store key row references
ROW_EDGE = EDGE_START + 2
ROW_WINRATE = ROW_EDGE + 1
ROW_ODDS = ROW_EDGE + 2
ROW_BETS_EVENT = ROW_EDGE + 3
ROW_EVENTS = ROW_EDGE + 4
ROW_VIG = ROW_EDGE + 5
ROW_UTIL = ROW_EDGE + 6

# ════════════════════════════════════════
# SHEET 2: SCENARIO OUTPUT
# ════════════════════════════════════════
ws2 = wb.create_sheet("Scenario Output")
ws2.sheet_properties.tabColor = "2E75B6"

for c in range(1, 9):
    ws2.column_dimensions[get_column_letter(c)].width = [2, 38, 18, 18, 18, 2, 18, 2][c-1]

ws2.merge_cells('B1:E1')
ws2['B1'] = "SCENARIO OUTPUT — ANNUAL P&L ESTIMATE"
ws2['B1'].font = TITLE_FONT
ws2['B1'].fill = dark_fill
style_range(ws2, 1, 2, 5, fill=dark_fill)

ws2.merge_cells('B2:E2')
ws2['B2'] = "Assumes $500/event/platform max bet, 2026 WSL CT season"
ws2['B2'].font = Font(name="Arial", italic=True, color="A0A0A0", size=9)
ws2['B2'].fill = dark_fill
style_range(ws2, 2, 2, 5, fill=dark_fill)

# ── CAPITAL DEPLOYED ──
r = 4
ws2.merge_cells(f'B{r}:E{r}')
ws2[f'B{r}'] = "CAPITAL DEPLOYED"
style_range(ws2, r, 2, 5, font=HEADER_FONT, fill=accent_fill)

r = 5
for i, h in enumerate(["Metric", "Low", "Medium", "High"]):
    ws2.cell(row=r, column=2+i, value=h).font = Font(name="Arial", bold=True, size=9, color="333333")
    ws2.cell(row=r, column=2+i).fill = light_fill
    ws2.cell(row=r, column=2+i).border = thin_border

A = "Assumptions"

calc_rows_capital = [
    ("Platforms used", f'=ROUND({A}!C{PLAT_LAST+1}*{A}!C{ROW_UTIL},0)',
     f'=ROUND({A}!C{PLAT_LAST+1}*{A}!D{ROW_UTIL},0)',
     f'=ROUND({A}!C{PLAT_LAST+1}*{A}!E{ROW_UTIL},0)', NUM_FMT),
    ("Bet size per platform per event", f'={A}!C6', f'={A}!C6', f'={A}!C6', DOLLAR_FMT),
    ("Bets per event (total across platforms)", f'={A}!C{ROW_BETS_EVENT}', f'={A}!D{ROW_BETS_EVENT}', f'={A}!E{ROW_BETS_EVENT}', NUM_FMT),
    ("Events bet on", f'={A}!C{ROW_EVENTS}', f'={A}!D{ROW_EVENTS}', f'={A}!E{ROW_EVENTS}', NUM_FMT),
    ("Total bets per year", f'=C8*C9', f'=D8*D9', f'=E8*E9', NUM_FMT),
    ("Total capital wagered per year", f'=C7*C10', f'=D7*D10', f'=E7*E10', DOLLAR_FMT),
]

for i, (label, low, mid, high, fmt) in enumerate(calc_rows_capital):
    r = 6 + i
    ws2.cell(row=r, column=2, value=label).font = LABEL_FONT
    ws2.cell(row=r, column=2).border = thin_border
    for j, val in enumerate([low, mid, high]):
        cell = ws2.cell(row=r, column=3+j, value=val)
        cell.font = FORMULA_FONT
        cell.number_format = fmt
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

# ── EXPECTED RETURNS ──
r = 13
ws2.merge_cells(f'B{r}:E{r}')
ws2[f'B{r}'] = "EXPECTED RETURNS (ANNUAL)"
style_range(ws2, r, 2, 5, font=HEADER_FONT, fill=accent_fill)

r = 14
for i, h in enumerate(["Metric", "Low", "Medium", "High"]):
    ws2.cell(row=r, column=2+i, value=h).font = Font(name="Arial", bold=True, size=9, color="333333")
    ws2.cell(row=r, column=2+i).fill = light_fill
    ws2.cell(row=r, column=2+i).border = thin_border

# Row refs from capital section
# Row 6=platforms, 7=bet size, 8=bets/event, 9=events, 10=total bets, 11=capital wagered

calc_rows_returns = [
    ("Average odds (decimal)", f'={A}!C{ROW_ODDS}', f'={A}!D{ROW_ODDS}', f'={A}!E{ROW_ODDS}', '0.00'),
    ("Win rate", f'={A}!C{ROW_WINRATE}', f'={A}!D{ROW_WINRATE}', f'={A}!E{ROW_WINRATE}', PCT_FMT),
    ("Expected value per $1 bet (before vig)", f'=(C15*C16)-1', f'=(D15*D16)-1', f'=(E15*E16)-1', '0.0%'),
    ("Vig drag per $1 bet", f'={A}!C{ROW_VIG}', f'={A}!D{ROW_VIG}', f'={A}!E{ROW_VIG}', PCT_FMT),
    ("Net EV per $1 bet (after vig)", f'=C17-C18', f'=D17-D18', f'=E17-E18', '0.0%'),
    ("Gross winnings (wins × payout)", f'=C10*C16*C7*C15', f'=D10*D16*D7*D15', f'=E10*E16*E7*E15', DOLLAR_FMT),
    ("Total wagered", f'=C11', f'=D11', f'=E11', DOLLAR_FMT),
    ("Net profit (Gross - Wagered)", f'=C20-C21', f'=D20-D21', f'=E20-E21', DOLLAR_FMT_NEG),
    ("ROI (Net / Wagered)", f'=IFERROR(C22/C21,0)', f'=IFERROR(D22/D21,0)', f'=IFERROR(E22/E21,0)', PCT_FMT),
]

for i, (label, low, mid, high, fmt) in enumerate(calc_rows_returns):
    r = 15 + i
    ws2.cell(row=r, column=2, value=label).font = LABEL_FONT
    ws2.cell(row=r, column=2).border = thin_border
    is_profit_row = "Net profit" in label or "ROI" in label
    for j, val in enumerate([low, mid, high]):
        cell = ws2.cell(row=r, column=3+j, value=val)
        cell.font = Font(name="Arial", bold=is_profit_row, color="000000", size=10 if not is_profit_row else 11)
        cell.number_format = fmt
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    if is_profit_row:
        for c in range(2, 6):
            ws2.cell(row=r, column=c).fill = PatternFill('solid', fgColor='E8F5E9' if "profit" in label.lower() else 'FFF3E0')

# ── MULTI-YEAR PROJECTION ──
r = 25
ws2.merge_cells(f'B{r}:E{r}')
ws2[f'B{r}'] = "MULTI-YEAR PROJECTION (Medium Scenario)"
style_range(ws2, r, 2, 5, font=HEADER_FONT, fill=accent_fill)

r = 26
for i, h in enumerate(["Year", "Cumulative Wagered", "Cumulative Profit", "Cumulative ROI"]):
    ws2.cell(row=r, column=2+i, value=h).font = Font(name="Arial", bold=True, size=9, color="333333")
    ws2.cell(row=r, column=2+i).fill = light_fill
    ws2.cell(row=r, column=2+i).border = thin_border

for yr in range(1, 6):
    r = 26 + yr
    ws2.cell(row=r, column=2, value=f"Year {yr}").font = LABEL_FONT
    ws2.cell(row=r, column=2).border = thin_border
    ws2.cell(row=r, column=3, value=f'=D21*{yr}').font = FORMULA_FONT
    ws2.cell(row=r, column=3).number_format = DOLLAR_FMT
    ws2.cell(row=r, column=3).border = thin_border
    ws2.cell(row=r, column=4, value=f'=D22*{yr}').font = FORMULA_FONT
    ws2.cell(row=r, column=4).number_format = DOLLAR_FMT_NEG
    ws2.cell(row=r, column=4).border = thin_border
    ws2.cell(row=r, column=5, value=f'=IFERROR(D{r}/C{r},0)').font = FORMULA_FONT
    ws2.cell(row=r, column=5).number_format = PCT_FMT
    ws2.cell(row=r, column=5).border = thin_border

# ── REALITY CHECK ──
r = 33
ws2.merge_cells(f'B{r}:E{r}')
ws2[f'B{r}'] = "REALITY CHECK"
style_range(ws2, r, 2, 5, font=HEADER_FONT, fill=PatternFill('solid', fgColor='C0392B'))

caveats = [
    "These numbers assume consistent edge over time — edge decay is real and unmodeled here.",
    "Getting limited on platforms will reduce effective platform count over time.",
    "The 'High' scenario (10% edge, 60% win rate at 3.0 avg odds) is extremely optimistic.",
    "The 'Low' scenario is the most honest starting expectation. Medium requires proven edge.",
    "Variance is enormous — a 55% win rate means ~45% of bets lose. Drawdowns will happen.",
    "No exchange markets exist for surf — sportsbook limitation risk is the binding constraint.",
    "This model does NOT account for bankroll management / Kelly criterion sizing.",
]

for i, caveat in enumerate(caveats):
    r = 34 + i
    ws2.cell(row=r, column=2, value=f"• {caveat}").font = Font(name="Arial", size=9, italic=True, color="8B0000")
    ws2.merge_cells(f'B{r}:E{r}')

# ── BREAKEVEN ANALYSIS ──
r = 42
ws2.merge_cells(f'B{r}:E{r}')
ws2[f'B{r}'] = "BREAKEVEN ANALYSIS"
style_range(ws2, r, 2, 5, font=HEADER_FONT, fill=accent_fill)

r = 43
for i, h in enumerate(["At These Avg Odds", "Breakeven Win Rate (incl. vig)", "Your Edge If Win Rate = 55%", "Profit at 55% / $500 / 60 bets"]):
    ws2.cell(row=r, column=2+i, value=h).font = Font(name="Arial", bold=True, size=9, color="333333")
    ws2.cell(row=r, column=2+i).fill = light_fill
    ws2.cell(row=r, column=2+i).border = thin_border

odds_vals = [1.91, 2.0, 2.5, 3.0, 4.0, 5.0]
for i, odds in enumerate(odds_vals):
    r = 44 + i
    ws2.cell(row=r, column=2, value=odds).font = LABEL_FONT
    ws2.cell(row=r, column=2).number_format = '0.00'
    ws2.cell(row=r, column=2).border = thin_border
    # Breakeven = 1/odds + vig_adjustment ≈ 1/odds * (1 + vig)
    ws2.cell(row=r, column=3, value=f'=(1/B{r})*(1+{A}!C{ROW_VIG})').font = FORMULA_FONT
    ws2.cell(row=r, column=3).number_format = PCT_FMT
    ws2.cell(row=r, column=3).border = thin_border
    # Edge if 55% win rate
    ws2.cell(row=r, column=4, value=f'=0.55-C{r}').font = FORMULA_FONT
    ws2.cell(row=r, column=4).number_format = '0.0%'
    ws2.cell(row=r, column=4).border = thin_border
    # Profit at 55%, $500, 60 bets/yr
    ws2.cell(row=r, column=5, value=f'=(0.55*B{r}*500*60)-(500*60)').font = FORMULA_FONT
    ws2.cell(row=r, column=5).number_format = DOLLAR_FMT_NEG
    ws2.cell(row=r, column=5).border = thin_border

# ── REAL ODDS EXAMPLES ──
ws3 = wb.create_sheet("Real Odds Examples")
ws3.sheet_properties.tabColor = "27AE60"

for c in range(1, 9):
    ws3.column_dimensions[get_column_letter(c)].width = [2, 30, 20, 14, 14, 14, 14, 2][c-1]

ws3.merge_cells('B1:G1')
ws3['B1'] = "REAL ODDS FROM NXTBETS DATA (2024-2025 CT Events)"
ws3['B1'].font = TITLE_FONT
ws3['B1'].fill = dark_fill
style_range(ws3, 1, 2, 7, fill=dark_fill)

r = 3
ws3.merge_cells(f'B{r}:G{r}')
ws3[f'B{r}'] = "OUTRIGHT WINNER ODDS — 2024 Fiji Pro (Men's CT)"
style_range(ws3, r, 2, 7, font=HEADER_FONT, fill=accent_fill)

r = 4
for i, h in enumerate(["Surfer", "Moneyline", "Decimal Odds", "Implied Prob", "If $500 Bet Wins", "Profit if Wins"]):
    ws3.cell(row=r, column=2+i, value=h).font = Font(name="Arial", bold=True, size=9, color="333333")
    ws3.cell(row=r, column=2+i).fill = light_fill
    ws3.cell(row=r, column=2+i).border = thin_border

fiji_odds = [
    ("John John Florence", 427, "=(B5/100)+1", "=1/C5", "=C5*500", "=E5-500"),
    ("Jack Robinson", 546, "=(B6/100)+1", "=1/C6", "=C6*500", "=E6-500"),
    ("Gabriel Medina", 611, "=(B7/100)+1", "=1/C7", "=C7*500", "=E7-500"),
    ("Griffin Colapinto", 911, "=(B8/100)+1", "=1/C8", "=C8*500", "=E8-500"),
    ("Ethan Ewing", 1182, "=(B9/100)+1", "=1/C9", "=C9*500", "=E9-500"),
    ("Italo Ferreira", 1240, "=(B10/100)+1", "=1/C10", "=C10*500", "=E10-500"),
]

for i, (name, ml, dec, prob, payout, profit) in enumerate(fiji_odds):
    r = 5 + i
    ws3.cell(row=r, column=2, value=name).font = LABEL_FONT
    ws3.cell(row=r, column=2).border = thin_border
    ws3.cell(row=r, column=3, value=ml).font = INPUT_FONT
    ws3.cell(row=r, column=3).number_format = '+#,##0'
    ws3.cell(row=r, column=3).border = thin_border
    ws3.cell(row=r, column=4, value=dec).font = FORMULA_FONT
    ws3.cell(row=r, column=4).number_format = '0.00'
    ws3.cell(row=r, column=4).border = thin_border
    ws3.cell(row=r, column=5, value=prob).font = FORMULA_FONT
    ws3.cell(row=r, column=5).number_format = PCT_FMT
    ws3.cell(row=r, column=5).border = thin_border
    ws3.cell(row=r, column=6, value=payout).font = FORMULA_FONT
    ws3.cell(row=r, column=6).number_format = DOLLAR_FMT
    ws3.cell(row=r, column=6).border = thin_border
    ws3.cell(row=r, column=7, value=profit).font = Font(name="Arial", bold=True, size=10)
    ws3.cell(row=r, column=7).number_format = DOLLAR_FMT
    ws3.cell(row=r, column=7).border = thin_border

# ── Pipe Pro 2025 ──
r = 13
ws3.merge_cells(f'B{r}:G{r}')
ws3[f'B{r}'] = "OUTRIGHT WINNER ODDS — 2025 Pipe Pro (Men's CT)"
style_range(ws3, r, 2, 7, font=HEADER_FONT, fill=accent_fill)

r = 14
for i, h in enumerate(["Surfer", "Moneyline", "Decimal Odds", "Implied Prob", "If $500 Bet Wins", "Profit if Wins"]):
    ws3.cell(row=r, column=2+i, value=h).font = Font(name="Arial", bold=True, size=9, color="333333")
    ws3.cell(row=r, column=2+i).fill = light_fill
    ws3.cell(row=r, column=2+i).border = thin_border

pipe_odds = [
    ("Jack Robinson", 544, "=(B15/100)+1", "=1/C15", "=C15*500", "=E15-500"),
    ("Filipe Toledo", 730, "=(B16/100)+1", "=1/C16", "=C16*500", "=E16-500"),
    ("Griffin Colapinto", 845, "=(B17/100)+1", "=1/C17", "=C17*500", "=E17-500"),
    ("Ethan Ewing", 918, "=(B18/100)+1", "=1/C18", "=C18*500", "=E18-500"),
    ("Italo Ferreira", 942, "=(B19/100)+1", "=1/C19", "=C19*500", "=E19-500"),
    ("Yago Dora", 1477, "=(B20/100)+1", "=1/C20", "=C20*500", "=E20-500"),
]

for i, (name, ml, dec, prob, payout, profit) in enumerate(pipe_odds):
    r = 15 + i
    ws3.cell(row=r, column=2, value=name).font = LABEL_FONT
    ws3.cell(row=r, column=2).border = thin_border
    ws3.cell(row=r, column=3, value=ml).font = INPUT_FONT
    ws3.cell(row=r, column=3).number_format = '+#,##0'
    ws3.cell(row=r, column=3).border = thin_border
    ws3.cell(row=r, column=4, value=dec).font = FORMULA_FONT
    ws3.cell(row=r, column=4).number_format = '0.00'
    ws3.cell(row=r, column=4).border = thin_border
    ws3.cell(row=r, column=5, value=prob).font = FORMULA_FONT
    ws3.cell(row=r, column=5).number_format = PCT_FMT
    ws3.cell(row=r, column=5).border = thin_border
    ws3.cell(row=r, column=6, value=payout).font = FORMULA_FONT
    ws3.cell(row=r, column=6).number_format = DOLLAR_FMT
    ws3.cell(row=r, column=6).border = thin_border
    ws3.cell(row=r, column=7, value=profit).font = Font(name="Arial", bold=True, size=10)
    ws3.cell(row=r, column=7).number_format = DOLLAR_FMT
    ws3.cell(row=r, column=7).border = thin_border

# Key insight row
r = 22
ws3.merge_cells(f'B{r}:G{r}')
ws3[f'B{r}'] = "KEY INSIGHT: At these odds, a single correct outright winner bet at +900 on a $500 stake returns $4,500+ profit."
ws3[f'B{r}'].font = Font(name="Arial", bold=True, size=10, color="27AE60")

r = 23
ws3.merge_cells(f'B{r}:G{r}')
ws3[f'B{r}'] = "But outright winners are rare (~10-15% implied probability for favorites). The edge must come from systematically identifying mispriced underdogs."
ws3[f'B{r}'].font = Font(name="Arial", italic=True, size=9, color="666666")

wb.save('/sessions/ecstatic-admiring-dijkstra/mnt/cowabunga/surf_betting_scenario_model.xlsx')
print("Done")
